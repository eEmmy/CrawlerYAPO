# -*- coding: utf-8 -*-
from collections import defaultdict
from bs4 import BeautifulSoup
import requests as requests
import xlsxwriter
import openpyxl
import sys

# Pega dados de um documento XLSX
def getXLSXdata():
	# Carrega o documento
	doc = openpyxl.load_workbook(filename="input/input.xlsx")

	# Salva os parametros das buscas
	searchs = []

	for data in doc['Planilha1'].iter_rows(values_only=True):
		searchs.append(data)

	return searchs

# Recupera todos os anuncios dá página
def getAdds(url):
	# Executa a requisição
	request = requests.get(url)

	# Verifica se a solicitação foi concluida com sucesso
	if request.status_code == 200:
		content = request.content
	else:
		return False

	# Pega todo o html da página
	soup = BeautifulSoup(content, 'html.parser')

	# Isola apenas as tabelas dos anuncios
	table = soup.find(name='table', attrs={'class':'listing_thumbs'})

	# Isola apenas as linhas da tabela
	rows = table.findAll(name='tr')

	# Guarda o objeto de retorno
	data = {}


	# Controlador de indices
	i = 0

	# Loop em rows
	for row in rows:
		data[i] = {
			'name': row.find(name='a', attrs={'class':'title'}).text.strip(),
			'img': row.findAll(name='img')[0]["src"],
			'price': row.find(name='span', attrs={'class':'price'}).text.strip(),
			'region': row.find(name='span', attrs={'class':'region'}).text.strip(),
			'link': row.find(name='a', attrs={'class':'title'})["href"]
		}

		i += 1
	
	return data

# Conta o numero de páginas da pesquisa
def countPages(url):
	# Executa a requisição
	request = requests.get(url)

	# Verifica se a solicitação foi concluida com sucesso
	if request.status_code == 200:
		content = request.content
	else:
		return False

	# Pega todo o html da página
	soup = BeautifulSoup(content, 'html.parser')

	# Isola o link para a ultima página
	linkContainer = soup.find(name='span', attrs={'class':'nohistory FloatRight'})

	# Recupera apenas o link
	link = linkContainer.find(name='a')['href']

	# Guarda a posição do numero da ultima página
	numPos = link.find("&o=")+3

	# Retorna o numero de páginas
	return int(link[numPos : 100000])

# Pega o nome do vendedor
def getSellerName(url):
	# Executa a requisição
	request = requests.get(url)

	# Verifica se a solicitação foi concluida com sucesso
	if request.status_code == 200:
		content = request.content
	else:
		return False

	# Pega todo o html da página
	soup = BeautifulSoup(content, 'html.parser')

	# Isola o link para a ultima página
	sellerName = soup.find(name='seller-info')['username']

	# Retorna o nome do vendedor
	return sellerName


def generateOutputFile(filename, content):
	workbook = xlsxwriter.Workbook(filename)
	worksheet = workbook.add_worksheet()

	rows = [
		['ID do produto', 'Nome', 'Link', 'Loja', 'Imagem', 'Preço', 'Vendedor', 'Região', 'Estado', 'Estoque inicial', 'Estoque atual', 'Estoque vendido']
	]

	i = 0

	for row in rows:
		for colName in row:
			worksheet.write(0, i, colName)
			i += 1

	rowNum = 1

	for page in content:
		for item in content[page]:
			# Escreve no documento
			worksheet.write(rowNum, 0, '')
			worksheet.write(rowNum, 1, content[page][item]['name'])
			worksheet.write(rowNum, 2, content[page][item]['link'])
			worksheet.write(rowNum, 3, 'YAPO')
			worksheet.write(rowNum, 4, content[page][item]['img'])
			worksheet.write(rowNum, 5, content[page][item]['price'])
			worksheet.write(rowNum, 6, content[page][item]['sellerName'])
			worksheet.write(rowNum, 7, content[page][item]['region'])
			worksheet.write(rowNum, 8, ''),
			worksheet.write(rowNum, 9, ''),
			worksheet.write(rowNum, 10, ''),
			worksheet.write(rowNum, 11, '')
			rowNum += 1
			pass
		pass

	workbook.close()

# Recupera as pesquisas a serem feitas
searchs = getXLSXdata()

# Dados das pesquisas
data = defaultdict()

searchs.pop(0)


searchs.pop(-1)

print("Recuperando anuncios.\n")

outputFilename = []

# Loop em searchs
for i in range(0, len(searchs)):
	# Armazena os parametros
	searchQuery = searchs[i][0].replace(' ', '+')
	outputFilename.append(searchs[i][-1])

	# Guarda a quantidade de páginas
	pagesAmount = countPages("https://www.yapo.cl/chile/todos_los_avisos?q=" + searchQuery)
	
	for e in range(1, pagesAmount+1):
		# Guarda a chave a ser adicionada a data
		toAdd = {
			e: getAdds("https://www.yapo.cl/chile/todos_los_avisos?q=" + searchQuery + "&o=" + str(e))
		}

		# Salva a busca
		data.setdefault(i, {})
		data[i].setdefault(e, toAdd)


# Pega apenas os dados, ignorando o None de defaultdict
data = data[0]

# Contadores de progresso
count = 0
counter = 1

print("Extraindo dados.")

for search in data:
	for page in data[search]:
		for page in data[search]:
			count += len(data[search][page])

# Loop em data
for search in data:
	# Loop em data[search]
	for page in data[search]:
		# Loop em data[search][page] (Pegar nome do vendedor)
		for item in data[search][page]:
			# Pega o nome do vendedor
			data[search][page][item] = {
				'name': data[search][page][item]['name'],
				'link': data[search][page][item]['link'],
				'sellerName': getSellerName(data[search][page][item]['link']),
				'img': data[search][page][item]['img'],
				'price': data[search][page][item]['price'],
				'region': data[search][page][item]['region'],
			}
			
			# Exibe o progresso
			print(str(counter) + "/" + str(count))
			
			# Aumenta o contador de progresso
			counter += 1
			pass
		pass
	pass

for search in data:
	try:
		generateOutputFile('output/'+outputFilename[search-1], data[search])
		print("Arquivo " + str(search) + " gerado com sucesso.")
	except Exception as e:
		continue